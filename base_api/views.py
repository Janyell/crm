#!/usr/bin/python
# -*- coding: utf-8 -*-
import json
import pdfkit
from base_api.full_views.analyze_debtors import full_analyze_debtors
from base_api.full_views.attach import *
from base_api.full_views.company_views import *
from base_api.full_views.role_views import *
from base_api.full_views.client_views import *
from base_api.full_views.order_views import *
from base_api.full_views.analyze_products import *
from base_api.full_views.analyze_managers import *
from base_api.full_views.analyze_sales_by_managers import *
from base_api.full_views.analyze_total_sales import *
from base_api.full_views.analyze_period import *
from base_api.full_views.product_views import *
from base_api.full_views.claim_views import *
# for excel
import openpyxl
import os
from django.http import HttpResponseRedirect
from django.template import Template, Context
from api.settings import MEDIA_ROOT, BASE_DIR
from base_api.models import *
from base_api.constants import *


def add_edit_role(request):
    return full_add_edit_role(request)


def delete_role(request):
    return full_delete_roles(request)


def get_roles(request):
    return full_get_roles(request)


def add_edit_company(request):
    return full_add_edit_company(request)


def delete_company(request):
    return full_delete_company(request)


def get_companies(request):
    return full_get_companies(request)


def add_edit_client(request):
    return full_add_edit_client(request)


def delete_client(request):
    return full_delete_clients(request)


def get_clients(request):
    return full_get_clients(request)


def add_edit_order(request):
    return full_add_edit_order(request)


def delete_order(request):
    return full_delete_order(request)


def get_orders(request):
    return full_get_orders(request)


def edit_order_for_factory(request):
    return full_edit_order_for_factory(request)


def analyst(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role != 0:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
        out.update({'page_title': "Аналитика"})
    return render(request, 'analyst/index.html', out)


def log_in(request):
    if not request.user.is_active:
        out = {}
        if request.method == 'POST':
            form = LoginForm(request.POST)
            username = request.POST["username"]
            password = request.POST["password"]
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                if Roles.objects.get(id=request.user.id).role != 0:
                    return HttpResponseRedirect('/orders/')
                else:
                    return HttpResponseRedirect('/')
            else:
                out.update({"error": 1})
        else:
            form = LoginForm()
        out.update({'login_form': form})
        out.update({'page_title': "Авторизация"})
        return render(request, 'log_in.html', out)
    else:
        if Roles.objects.get(id=request.user.id).role != 0:
            return HttpResponseRedirect('/orders/')
        else:
            return HttpResponseRedirect('/')


def log_out(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    logout(request)
    return HttpResponseRedirect("/login/")


def page_not_found(request):
    out = {}
    try:
        user_role = Roles.objects.get(id=request.user.id).role
        out = {'user_role': user_role}
        out.update({'page_title': "Страница не найдена"})
        return render(request, 'page_not_found.html', out)
    except ObjectDoesNotExist:
        return HttpResponseRedirect("/login/")


def permission_deny(request):
    out = {}
    user_role = Roles.objects.get(id=request.user.id).role
    out.update({'user_role': user_role})
    out.update({'page_title': "К сожалению, у вас нет доступа к данной странице"})
    return render(request, 'page_not_found.html', out)


def get_old_orders(request):
    return full_get_old_orders(request)


def add_in_archive(request):
    return full_add_in_archive(request)


def get_interested_clients(request):
    return full_get_interested_clients(request)


def analyze_products(request):
    return full_analyze_products(request)


def view_analyzed_product(request):
    return full_view_analyzed_product(request)


def analyze_managers(request):
    return full_analyze_managers(request)


def analyze_total_sales(request):
    return full_analyze_total_sales(request)


def analyze_sales_by_managers(request):
    return full_analyze_sales_by_managers(request)


def analyze_period(request):
    return full_analyze_period(request)


def analyze_period_product_groups(request):
    return full_analyze_period_product_groups(request)


def json_response(func):
    """
    A decorator thats takes a view response and turns it
    into json. If a callback is added through GET or POST
    the response is JSONP.
    """

    def decorator(request, *args, **kwargs):
        objects = func(request, *args, **kwargs)
        if isinstance(objects, HttpResponse):
            return objects
        try:
            data = json.dumps(objects)
            if 'callback' in request.REQUEST:
                # a jsonp response!
                data = '%s(%s);' % ('func', data)
                return HttpResponse(data, "text/javascript")
        except:
            data = json.dumps(str(objects))
        return HttpResponse(data, "application/json")

    return decorator


@json_response
def give_order_status(request):
    order_unic = request.GET['id']
    result = {}
    try:
        order = Orders.objects.get(unique_number=order_unic)
        result['order_status'] = order.order_status
        result['order_date'] = str(order.order_date)
        result['payment_date'] = str(order.payment_date)
        result['bill_status'] = order.bill_status
        result['ready_date'] = str(order.ready_date)
    except ObjectDoesNotExist:
        result['error'] = u'order_is_not_exist'
    return result


def get_products(request):
    return full_get_products(request)


def delete_product(request):
    return full_delete_product(request)


def edit_order_for_other_managers(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role == 2:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
    if request.method == 'POST':
        if 'pk' in request.POST:
            pk = request.POST['pk']
            comment = request.POST['comment']
            factory_comment = None
            if 'factory_comment' in request.POST:
                factory_comment = request.POST['factory_comment']
            new_order = Orders.objects.get(id=pk, is_deleted=0)
            is_comment_my = False
            if new_order.role_id == request.user.id:
                is_comment_my = True
            new_order.is_comment_my = is_comment_my
            new_order.comment = comment
            new_order.factory_comment = factory_comment
            new_order.save(force_update=True)
            get_params = '?'
            if 'search' in request.GET:
                search = request.GET.get('search')
                get_params += 'search=' + unicode(search)
                return HttpResponseRedirect('/search/' + get_params)
            get_params += get_request_param_as_string(request)
            return HttpResponseRedirect('/orders/' + get_params)
        else:
            get_params = '?'
            if 'search' in request.GET:
                search = request.GET.get('search')
                get_params += 'search=' + unicode(search)
                return HttpResponseRedirect('/search/' + get_params)
            get_params += get_request_param_as_string(request)
            return HttpResponseRedirect('/orders/' + get_params)
    else:
        if 'id' in request.GET:
            id_order = request.GET['id']
            out.update({"error": 0})
            order = Orders.objects.get(pk=id_order, is_deleted=0)
            form = OrdersForm({'comment': order.comment, 'factory_comment': order.factory_comment})
            out.update({'order_form': form})
            out.update({'page_title': "Редактирование заказа"})
        else:
            get_params = '?'
            if 'search' in request.GET:
                search = request.GET.get('search')
                get_params += 'search=' + unicode(search)
                return HttpResponseRedirect('/search/' + get_params)
            get_params += get_request_param_as_string(request)
            return HttpResponseRedirect('/orders/' + get_params)
    return render(request, 'order_claim/edit_order_for_other_managers.html', out)


def edit_claim_for_other_managers(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role == 2:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
    if request.method == 'POST':
        if 'pk' in request.POST:
            pk = request.POST['pk']
            comment = request.POST['comment']
            factory_comment = None
            if 'factory_comment' in request.POST:
                factory_comment = request.POST['factory_comment']
            new_order = Orders.objects.get(id=pk, is_deleted=0, is_claim=1)
            new_order.comment = comment
            new_order.factory_comment = factory_comment
            is_comment_my = False
            if new_order.role_id == request.user.id:
                is_comment_my = True
            new_order.is_comment_my = is_comment_my
            new_order.save(force_update=True)
            get_params = '?'
            if 'search' in request.GET:
                search = request.GET.get('search')
                get_params += 'search=' + unicode(search)
                return HttpResponseRedirect('/search/' + get_params)
            get_params += get_request_param_as_string(request)
            return HttpResponseRedirect('/claims/' + get_params)
        else:
            get_params = '?'
            if 'search' in request.GET:
                search = request.GET.get('search')
                get_params += 'search=' + unicode(search)
                return HttpResponseRedirect('/search/' + get_params)
            get_params += get_request_param_as_string(request)
            return HttpResponseRedirect('/claims/' + get_params)
    else:
        if 'id' in request.GET:
            id_order = request.GET['id']
            out.update({"error": 0})
            order = Orders.objects.get(pk=id_order, is_deleted=0, is_claim=1)
            form = OrdersForm({'comment': order.comment, 'factory_comment': order.factory_comment})
            out.update({'order_form': form})
            out.update({'page_title': "Редактирование заявки"})
        else:
            get_params = '?'
            if 'search' in request.GET:
                search = request.GET.get('search')
                get_params += 'search=' + unicode(search)
                return HttpResponseRedirect('/search/' + get_params)
            get_params += get_request_param_as_string(request)
            return HttpResponseRedirect('/claims/' + get_params)
    return render(request, 'order_claim/edit_order_for_other_managers.html', out)


def get_claims(request):
    return full_get_claims(request)


def add_edit_claim(request):
    return full_add_edit_claim(request)


def delete_claim(request):
    return full_delete_claim(request)


def delete_from_archive(request):
    return full_delete_from_archive(request)


def upload_order_files(request):
    return upload_order_file(request)


def delete_order_files(request):
    return delete_order_file(request)


def upload_client_files(request):
    return upload_client_file(request)


def delete_client_files(request):
    return delete_client_file(request)


def fix_bd_org_type(request):
    clients = Clients.objects.all()
    for client in clients:
        if u'ИП' in client.organization:
            client.organization_type = u'ИП'
            client.organization = client.organization[3:]
        elif u'ООО' in client.organization:
            client.organization_type = u'ООО'
            client.organization = client.organization[4:]
        elif u'ЗАО' in client.organization:
            client.organization_type = u'ЗАО'
            client.organization = client.organization[4:]
        elif u'ОАО' in client.organization:
            client.organization_type = u'ОАО'
            client.organization = client.organization[4:]
        elif u'НКО' in client.organization:
            client.organization_type = u'НКО'
            client.organization = client.organization[4:]
        elif u'ТСЖ' in client.organization:
            client.organization_type = u'ТСЖ'
            client.organization = client.organization[4:]
        elif u'ОП' in client.organization:
            client.organization_type = u'ОП'
            client.organization = client.organization[3:]
        client.save(update_fields=["organization", "organization_type"])
    return HttpResponseRedirect('/clients/interested/')


def made_excel(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role != 0:
        return HttpResponseRedirect('/oops/')

    random_filename = str(''.join(random.choice(string.lowercase) for i in range(10))) + '.xlsx'
    filename = MEDIA_ROOT + '/uploads/' + random_filename

    wb = openpyxl.Workbook()
    sheet = wb.active

    table_name = request.GET.get('table')
    table = EXCEL_TABLES.get(table_name)
    cols = request.POST.getlist('cols[]')

    if u'all' in cols:
        cols.remove(u'all')

    row_index = 1

    if table_name == u'products':
        table_objects = table.objects.filter(is_deleted=0).all()
        column_index = 1
        if u'group' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Группа"
            column_index = column_index + 1
        if u'title' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Название"
            column_index = column_index + 1
        if u'price' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Цена"
            column_index = column_index + 1
        if u'is_active' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Статус"

        for table_object in table_objects:
            row_index += 1
            column_index = 1
            for col in cols:
                if col == u'group':
                    if getattr(table_object, col):
                        sheet.cell(row=row_index, column=column_index).value = unicode(getattr(table_object, col).title)
                elif col == u'is_active':
                    if getattr(table_object, col):
                        sheet.cell(row=row_index, column=column_index).value = u'Активный'
                    else:
                        sheet.cell(row=row_index, column=column_index).value = u'Не активный'
                else:
                    sheet.cell(row=row_index, column=column_index).value = unicode(getattr(table_object, col))
                column_index += 1

    elif table_name == u'clients':
        table_objects = table.objects.filter(is_deleted=0).all()
        column_index = 1
        if u'organization' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Организация"
            column_index = column_index + 1
        if u'organization_phone' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Телефон организации"
            column_index = column_index + 1
        if u'name' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Контактное лицо"
            column_index = column_index + 1
        if u'person_phone' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Телефон контактного лица"
            column_index = column_index + 1
        if u'email' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Email"

        for table_object in table_objects:
            row_index += 1
            column_index = 1
            for col in cols:
                if col == u'name':
                    sheet.cell(row=row_index, column=column_index).value = \
                        unicode(getattr(table_object, 'last_name')) +\
                        ' ' + unicode(getattr(table_object, 'name')) +\
                        ' ' + unicode(getattr(table_object, 'patronymic'))
                else:
                    sheet.cell(row=row_index, column=column_index).value = unicode(getattr(table_object, col))
                column_index += 1

    elif table_name == u'claims':
        table_objects = table.objects.filter(is_deleted=0, is_claim=1).all()
        column_index = 1
        if u'role' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Менеджер"
            column_index = column_index + 1
        if u'order_date' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Дата заявки"
            column_index = column_index + 1
        if u'organization' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Название организации"
            column_index = column_index + 1
        if u'company' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Компания"
            column_index = column_index + 1
        if u'account_number' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Номер счета"
            column_index = column_index + 1
        if u'bill' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Сумма счета"
            column_index = column_index + 1
        if u'products' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Продукция"
            column_index = column_index + 1
        if u'order_status' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Статус"
            column_index = column_index + 1
        if u'comment' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Комментарии"

        for table_object in table_objects:
            row_index += 1
            column_index = 1
            for col in cols:
                if col == u'organization':
                    client = getattr(table_object, 'client')
                    if client.organization == '':
                        organization_or_full_name = client.last_name + ' ' + client.name + ' ' + client.patronymic
                    else:
                        organization_or_full_name = client.organization
                    sheet.cell(row=row_index, column=column_index).value = unicode(organization_or_full_name)
                elif col == u'company':
                    if getattr(table_object, col):
                        sheet.cell(row=row_index, column=column_index).value = unicode(getattr(table_object, col).title)
                elif col == u'products':
                    pass
                elif col == u'order_status':
                    bill_status = getattr(table_object, 'bill_status')
                    if bill_status == 0:
                        sheet.cell(row=row_index, column=column_index).value = 'Выставлен'
                    elif bill_status == 1:
                        sheet.cell(row=row_index, column=column_index).value = 'Нужна доплата'
                    elif bill_status == 2:
                        sheet.cell(row=row_index, column=column_index).value = 'Оплачен'
                    elif bill_status == 4:
                        sheet.cell(row=row_index, column=column_index).value = 'Устно'
                    elif bill_status == 5:
                        sheet.cell(row=row_index, column=column_index).value = 'Подбор'
                    else:
                        sheet.cell(row=row_index, column=column_index).value = ''
                else:
                    sheet.cell(row=row_index, column=column_index).value = unicode(getattr(table_object, col))
                column_index += 1

    elif table_name == u'orders':
        table_objects = table.objects.filter(is_deleted=0, is_claim=0).all()
        column_index = 1
        if u'role' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Менеджер"
            column_index = column_index + 1
        if u'organization' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Название организации"
            column_index = column_index + 1
        if u'unique_number' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Номер заказа"
            column_index = column_index + 1
        if u'company' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Компания"
            column_index = column_index + 1
        if u'account_number' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Номер счета"
            column_index = column_index + 1
        if u'bill' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Сумма счета"
            column_index = column_index + 1
        if u'payment_date' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Дата оплаты"
            column_index = column_index + 1
        if u'products' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Продукция"
            column_index = column_index + 1
        if u'order_status' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Статус"
            column_index = column_index + 1
        if u'ready_date' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Дата готовности"
            column_index = column_index + 1
        if u'city' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Доставка город"
            column_index = column_index + 1
        if u'transport_company' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Транспортная компания"
            column_index = column_index + 1
        if u'comment' in cols:
            sheet.cell(row=row_index, column=column_index).value = "Комментарии"

        for table_object in table_objects:
            row_index += 1
            column_index = 1
            for col in cols:
                if col == u'organization':
                    client = getattr(table_object, 'client')
                    if client.organization == '':
                        organization_or_full_name = client.last_name + ' ' + client.name + ' ' + client.patronymic
                    else:
                        organization_or_full_name = client.organization
                    sheet.cell(row=row_index, column=column_index).value = unicode(organization_or_full_name)
                elif col == u'company':
                    if getattr(table_object, col):
                        sheet.cell(row=row_index, column=column_index).value = unicode(getattr(table_object, col).title)
                elif col == u'transport_company':
                    if getattr(table_object, 'transport_campaign'):
                        sheet.cell(row=row_index, column=column_index).value = \
                            unicode(getattr(table_object, 'transport_campaign').title)
                elif col == u'city':
                    if getattr(table_object, 'transport_campaign'):
                        sheet.cell(row=row_index, column=column_index).value = unicode(getattr(table_object, col).name)
                elif col == u'products':
                    pass
                elif col == u'order_status':
                    order_status = getattr(table_object, 'order_status')
                    if order_status == 0:
                        sheet.cell(row=row_index, column=column_index).value = 'В производстве'
                    elif order_status == -1:
                        sheet.cell(row=row_index, column=column_index).value = 'Отгружен'
                    elif order_status == 2:
                        sheet.cell(row=row_index, column=column_index).value = 'Готов'
                    else:
                        sheet.cell(row=row_index, column=column_index).value = ''
                else:
                    sheet.cell(row=row_index, column=column_index).value = unicode(getattr(table_object, col))
                column_index += 1

    # сохраняем данные
    wb.save(filename)
    return HttpResponseRedirect('/media/uploads/' + random_filename)


def made_excel_products(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role != 0:
        return HttpResponseRedirect('/oops/')
    filename = MEDIA_ROOT + '/' + str('uploads/products.xlsx')
    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb['list']
    products = Products.objects.filter(is_deleted=0).all()
    row_index = 1
    sheet.cell(row=row_index, column=1).value = "Id"
    sheet.cell(row=row_index, column=2).value = "Название"
    for product in products:
        row_index += 1
        column_index = 1
        sheet.cell(row=row_index, column=column_index).value = product.id
        column_index += 1
        sheet.cell(row=row_index, column=column_index).value = product.title
    # сохраняем данные
    wb.save(filename)
    return HttpResponseRedirect('/')


def analyze_debtors(request):
    return full_analyze_debtors(request)


def edit_product(request):
    return full_edit_product(request)


def get_documents(request):
    return render(request, 'get_documents.html')


def search(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    if 'page' in request.GET and 'length' in request.GET:
        page = int(request.GET['page'])
        length = int(request.GET['length'])
        start = (page - 1) * length
        out.update({'start': start})
    user_role = Roles.objects.get(id=request.user.id).role
    out.update({'user_role': user_role})
    out.update({'user': Roles.objects.get(id=request.user.id)})
    out.update({'roles': Roles.objects.filter(is_deleted=0).filter(Q(role=1) | Q(role=0)).all()})
    if 'search' in request.GET:
        search_word = request.GET['search'] + '*'
    else:
        search_word = u'*'
    order_list = list(Orders.search.query(search_word).filter(is_deleted=0, is_claim=0, in_archive=0))
    all_clients = list(Clients.search.query(search_word))
    orders_from_fk = []
    for client in all_clients:
        for order in Orders.objects.filter(client=client.id, is_deleted=0, is_claim=0, in_archive=0).all():
            orders_from_fk.append(order)
    all_companies = list(Companies.search.query(search_word))
    for company in all_companies:
        for order in Orders.objects.filter(company=company.id, is_deleted=0, is_claim=0, in_archive=0).all():
            orders_from_fk.append(order)
    order_list += orders_from_fk
    for order in order_list:
        if order.client.organization == '':
            order.client.organization_or_full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        else:
            order.client.organization_or_full_name = order.client.organization
        order.client.full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        prs = Order_Product.objects.filter(order_id=order.id, is_deleted=0)
        products_list = []
        for pr in prs:
            products_list.append(pr)
        order.products = products_list
        if order.order_status == 0:
            order.order_status = 'В производстве'
        elif order.order_status == -1:
            order.order_status = 'Отгружен'
            if order.shipped_date is not None:
                order.is_shipped = 1
                order.shipped_date = date(order.shipped_date.year, order.shipped_date.month, order.shipped_date.day)
        elif order.order_status == 2:
            order.order_status = 'Готов'
        else:
            order.order_status = ''
        if order.bill is not None:
            order.bill_right_format = right_money_format(order.bill)
        order.brought_sum_right_format = 0
        order.debt_right_format = 0
        order.is_in_debt = False
        if order.bill_status == 2:
            order.is_full_pay = True
        else:
            order.is_full_pay = False
        if order.brought_sum is not None and order.bill is not None:
            if (order.bill - order.brought_sum) > 0:
                order.is_in_debt = True
            else:
                order.is_in_debt = False
            order.brought_sum_right_format = right_money_format(order.brought_sum)
            order.debt_right_format = right_money_format(int(order.bill) - int(order.brought_sum))
        if order.is_full_pay or order.bill_status == 3:
            order.is_in_debt = False
        order.files = []
        if Order_Files.objects.filter(order_id=order.id).all() is not None:
            for order_file in Order_Files.objects.filter(order_id=order.id).all():
                if order_file.file is not None and order_file.file != '':
                    order_file.name = order_file.title
                    order_file.url = order_file.file.url
                    order.files.append(order_file)

    archive_order_list = list(Orders.search.query(search_word).filter(is_deleted=0, is_claim=0, in_archive=1))
    all_clients = list(Clients.search.query(search_word))
    orders_from_fk = []
    for client in all_clients:
        for order in Orders.objects.filter(client=client.id, is_deleted=0, is_claim=0, in_archive=1).all():
            orders_from_fk.append(order)
    all_companies = list(Companies.search.query(search_word))
    for company in all_companies:
        for order in Orders.objects.filter(company=company.id, is_deleted=0, is_claim=0, in_archive=1).all():
            orders_from_fk.append(order)
    archive_order_list += orders_from_fk
    for order in archive_order_list:
        if order.client.organization == '':
            order.client.organization_or_full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        else:
            order.client.organization_or_full_name = order.client.organization
        order.client.full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        prs = Order_Product.objects.filter(order_id=order.id, is_deleted=0)
        products_list = []
        for pr in prs:
            products_list.append(pr)
        order.products = products_list
        if order.order_status == 0:
            order.order_status = 'В производстве'
        elif order.order_status == -1:
            order.order_status = 'Отгружен'
            if order.shipped_date is not None:
                order.is_shipped = 1
                order.shipped_date = date(order.shipped_date.year, order.shipped_date.month, order.shipped_date.day)
        elif order.order_status == 2:
            order.order_status = 'Готов'
        else:
            order.order_status = ''
        if order.bill is not None:
            order.bill_right_format = right_money_format(order.bill)
        order.brought_sum_right_format = 0
        order.debt_right_format = 0
        order.is_in_debt = False
        if order.bill_status == 2:
            order.is_full_pay = True
        else:
            order.is_full_pay = False
        if order.brought_sum is not None and order.bill is not None:
            if (order.bill - order.brought_sum) > 0:
                order.is_in_debt = True
            else:
                order.is_in_debt = False
            order.brought_sum_right_format = right_money_format(order.brought_sum)
            order.debt_right_format = right_money_format(int(order.bill) - int(order.brought_sum))
        if order.is_full_pay:
            order.is_in_debt = False
        order.files = []
        if Order_Files.objects.filter(order_id=order.id).all() is not None:
            for order_file in Order_Files.objects.filter(order_id=order.id).all():
                if order_file.file is not None and order_file.file != '':
                    order_file.name = order_file.title
                    order_file.url = order_file.file.url
                    order.files.append(order_file)

    claim_list = list(Orders.search.query(search_word).filter(is_deleted=0, is_claim=1, in_archive=0))
    all_clients = list(Clients.search.query(search_word))
    orders_from_fk = []
    for client in all_clients:
        for order in Orders.objects.filter(client=client.id, is_deleted=0, is_claim=1, in_archive=0).all():
            orders_from_fk.append(order)
    all_companies = list(Companies.search.query(search_word))
    for company in all_companies:
        for order in Orders.objects.filter(company=company.id, is_deleted=0, is_claim=1, in_archive=0).all():
            orders_from_fk.append(order)
    claim_list += orders_from_fk
    for order in claim_list:
        if order.client.organization == '':
            order.client.organization_or_full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        else:
            order.client.organization_or_full_name = order.client.organization
        order.client.full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        prs = Order_Product.objects.filter(order_id=order.id, is_deleted=0)
        products_list = []
        for pr in prs:
            products_list.append(pr)
        order.products = products_list
        if order.bill_status == 0:
            order.bill_status = 'Выставлен'
        elif order.bill_status == 1:
            order.bill_status = 'Нужна доплата'
        elif order.bill_status == 2:
            order.bill_status = 'Оплачен'
        elif order.bill_status == 4:
            order.bill_status = 'Устно'
        elif order.bill_status == 5:
            order.bill_status = 'Подбор'
        else:
            order.bill_status = ''
        if order.bill != None:
            orders_count_str = str(order.bill)
            orders_count_str_reverse = orders_count_str[::-1]
            orders_count_str_right_format = ''
            for i in range(0, len(orders_count_str) / 3 + 1):
                j = 3
                if i != (len(orders_count_str) / 3):
                    orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[i * j] + \
                                                    orders_count_str_reverse[i * j + 1] + orders_count_str_reverse[
                                                        i * j + 2] + ' '
                else:
                    if (len(orders_count_str) % 3) == 2:
                        orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[
                            i * j] + \
                                                        orders_count_str_reverse[i * j + 1]
                    elif (len(orders_count_str) % 3) == 1:
                        orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[i * j]
            orders_count_str = orders_count_str_right_format[::-1]
            order.bill = orders_count_str
        order.files = []
        if Order_Files.objects.filter(order_id=order.id).all() is not None:
            for order_file in Order_Files.objects.filter(order_id=order.id).all():
                if order_file.file:
                    order_file.name = order_file.title
                    order_file.url = order_file.file.url
                    order.files.append(order_file)

    client_list = list(Clients.search.query(search_word).filter(is_deleted=0, is_interested=0))
    for c in client_list:
        c.person_full_name = c.last_name + ' ' + c.name + ' ' + c.patronymic
        c.files = []
        if Client_Files.objects.filter(client_id=c.id).all() is not None:
            for client_file in Client_Files.objects.filter(client_id=c.id).all():
                if client_file.file is not None and client_file.file != '':
                    client_file.name = client_file.title
                    client_file.url = client_file.file.url
                    c.files.append(client_file)

    interested_client_list = list(Clients.search.query(search_word).filter(is_deleted=0, is_interested=1))
    for c in interested_client_list:
        c.person_full_name = c.last_name + ' ' + c.name + ' ' + c.patronymic
        c.files = []
        if Client_Files.objects.filter(client_id=c.id).all() is not None:
            for client_file in Client_Files.objects.filter(client_id=c.id).all():
                if client_file.file is not None and client_file.file != '':
                    client_file.name = client_file.title
                    client_file.url = client_file.file.url
                    c.files.append(client_file)
    if client_list or interested_client_list or order_list or archive_order_list or claim_list:
        out.update({'search_results': 1})
    out.update({'page_title': "Клиенты"})
    out.update({'clients': client_list})
    out.update({'interested_clients': interested_client_list})
    out.update({'orders': order_list})
    out.update({'old_orders': archive_order_list})
    out.update({'claims': claim_list})
    return render(request, 'search/search.html', out)


def fix_file_nodes(request):
    files = Order_Files.objects.all()
    for file in files:
        if file.title != 'Emails_of_clients.xlsx':
            save_file_in_node(file)
    files = Client_Files.objects.all()
    for file in files:
        save_file_in_node(file)
    return HttpResponseRedirect('/')


def fix_cities(request):
    # filename = MEDIA_ROOT + '/' + str('uploads/city.xlsx')
    # wb = openpyxl.load_workbook(filename=filename)
    # sheet = wb['list']
    # column_index = 4
    # for row_index in xrange(2, 10971):
    # city = Cities.objects.create(name=sheet.cell(row=row_index, column=column_index).value)
    #     row_index += 1

    # cities_list = list(Cities.objects.values_list('name', flat=True))
    # cities_list = [x.lower() for x in cities_list]
    # orders = Orders.objects.all()
    # for order in orders:
    #     if order.city_old:
    #         order.city_old = order.city_old.strip()
    #         if order.city_old.lower() not in cities_list:
    #             city = Cities.objects.create(name=order.city_old)
    #             cities_list.append(order.city_old.lower())
    #         else:
    #             city = Cities.objects.filter(name=order.city_old).first()
    #         order.city = city
    #         order.save(update_fields=["city"])
    # cities = Cities.objects.order_by('name').all()
    # for x in range(0, 287):
    #     if cities[x].name == cities[x+1].name:
    #         for order in Orders.objects.filter(city=cities[x+1]):
    #             order.city = cities[x]
    #             order.save()
    #         cities[x+1].delete()
    # options = {
    #     'page-size': 'Letter',
    #     'margin-top': '0.75in',
    #     'margin-right': '0.75in',
    #     'margin-bottom': '0.75in',
    #     'margin-left': '0.75in',
    #     'encoding': "UTF-8",
    #     'no-outline': None
    # }
    # pdfkit.from_file('/Users/megge/Documents/crm/templates/kp/kp.html',
    #                  '/Users/megge/Documents/crm/templates/kp/google.pdf',
    #                  options=options)
    # from wkhtmltopdf import WKhtmlToPdf
    # wkhtmltopdf.render()
    # wkhtmltopdf(url='/Users/megge/Documents/crm/templates/kp/kp.html', output_file='/Users/megge/Documents/crm/templates/kp/google.pdf')
    return render(request, 'kp/kp.html')


def get_product_groups(request):
    return full_get_product_groups(request)


def delete_product_group(request):
    return full_delete_product_group(request)


def edit_product_group(request):
    return full_edit_product_group(request)


def get_settings(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role != 0:
        return HttpResponseRedirect('/oops/')
    out.update({'user_role': user_role})
    return render(request, 'setting/get_settings.html', out)


def get_reports(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    out.update({'sources': Sources.objects.filter(is_deleted=0)})
    out.update({'transport_campaigns': TransportCampaigns.objects.filter(is_deleted=0)})
    if 'page' in request.GET and 'length' in request.GET:
        page = int(request.GET['page'])
        length = int(request.GET['length'])
        start = (page - 1) * length
        out.update({'start': start})
        out.update({'page': page})
        out.update({'length': length})
    user_role = Roles.objects.get(id=request.user.id).role
    out.update({'user_role': user_role})
    sort_key = request.GET.get('sort', DEFAULT_SORT_TYPE_FOR_ORDER)
    sort = SORT_TYPE_FOR_ORDER.get(sort_key, DEFAULT_SORT_TYPE_FOR_ORDER)
    orders = Orders.objects.filter(is_deleted=0, is_claim=0, role=request.user.id)
    if 'source' in request.GET:
        source = int(request.GET.get('source'))
        if source != -1:
            orders = orders.filter(source=source)
    if 'month-date' in request.GET:
        month_date = request.GET.get('month-date')
    else:
        month_date = date.today().strftime("%Y-%m")
    out.update({'month_date': month_date})
    orders = orders.filter(shipped_date__year=month_date[:4],
                           shipped_date__month=month_date[5:7])
    try:
        orders = orders.filter(in_archive=0).order_by(sort)
    except TypeError:
        orders = orders.filter(in_archive=0).order_by(*sort)
    shipped_sum = 0
    for order in orders:
        shipped_sum += order.bill
    out.update({'page_title': "Отчет"})
    out.update({'shipped_sum': shipped_sum})
    number = request.GET.get('length', DEFAULT_NUMBER_FOR_PAGE)
    orders_pages = Paginator(orders, number)
    page = request.GET.get('page')
    try:
        order_list = orders_pages.page(page)
    except PageNotAnInteger:
        order_list = orders_pages.page(1)
    except EmptyPage:
        order_list = orders_pages.page(orders_pages.num_pages)
    for order in order_list:
        if order.client.organization == '':
            order.client.organization_or_full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        else:
            order.client.organization_or_full_name = order.client.organization
        order.client.full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        prs = Order_Product.objects.filter(order_id=order.id, is_deleted=0)
        products_list = []
        for pr in prs:
            products_list.append(pr)
        order.products = products_list
        if order.order_status == 0:
            order.order_status = 'В производстве'
        elif order.order_status == -1:
            order.order_status = 'Отгружен'
            if order.shipped_date is not None:
                order.is_shipped = 1
                order.shipped_date = date(order.shipped_date.year, order.shipped_date.month, order.shipped_date.day)
        elif order.order_status == 2:
            order.order_status = 'Готов'
        else:
            order.order_status = ''
        if order.bill is not None:
            order.bill_right_format = right_money_format(order.bill)
        order.brought_sum_right_format = 0
        order.debt_right_format = 0
        order.is_in_debt = False
        if order.bill_status == 2:
            order.is_full_pay = True
        else:
            order.is_full_pay = False
        if order.brought_sum is not None and order.bill is not None:
            if (order.bill - order.brought_sum) > 0:
                order.is_in_debt = True
            else:
                order.is_in_debt = False
            order.brought_sum_right_format = right_money_format(order.brought_sum)
            order.debt_right_format = right_money_format(int(order.bill) - int(order.brought_sum))
        if order.is_full_pay or order.bill_status == 3:
            order.is_in_debt = False
        order.files = []
        if Order_Files.objects.filter(order_id=order.id).all() is not None:
            for order_file in Order_Files.objects.filter(order_id=order.id).all():
                if order_file.file is not None and order_file.file != '':
                    order_file.name = order_file.title
                    order_file.url = order_file.file.url
                    order.files.append(order_file)
    user_role = Roles.objects.get(id=request.user.id).role
    out.update({'user_role': user_role})
    out.update({'orders': order_list})
    out.update({'count': orders.count()})
    # IN_PRODUCTION status = 0
    out.update({'count_in_production': orders.filter(order_status=0).count()})
    return render(request, 'get_reports.html', out)


def analyze_cities(request):
    return render(request, 'analyst/analyze_cities.html')


def analyze_product_groups(request):
    return render(request, 'analyst/analyze_product_groups.html')


def view_analyzed_product_group(request):
    return render(request, 'analyst/view_analyzed_product.html')


def get_related_claims(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    out.update({'sources': Sources.objects.filter(is_deleted=0)})
    out.update({'transport_campaigns': TransportCampaigns.objects.filter(is_deleted=0)})
    out.update({'roles': Roles.objects.filter(is_deleted=0).filter(Q(role=1) | Q(role=0)).all()})
    if 'page' in request.GET and 'length' in request.GET:
        page = int(request.GET['page'])
        length = int(request.GET['length'])
        start = (page - 1) * length
        out.update({'start': start})
    user_id = request.user.id
    out.update({'user_id': user_id})
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role == 2:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
    sort_key = request.GET.get('sort', DEFAULT_SORT_TYPE_FOR_CLAIM)
    sort = SORT_TYPE_FOR_CLAIM.get(sort_key, DEFAULT_SORT_TYPE_FOR_CLAIM)
    claim = Orders.objects.get(id=request.GET['id'])
    orders = claim.related_orders.filter(is_deleted=0, is_claim=1, in_archive=0)
    if 'source' in request.GET:
        source = int(request.GET.get('source'))
        if source != -1:
            orders = orders.filter(source=source)
    if 'managers[]' in request.GET:
        managers = request.GET.getlist('managers[]')
        orders = orders.filter(role__in=managers)
        out.update({'managers': managers})
    try:
        orders = orders.order_by(sort)
    except TypeError:
        orders = orders.order_by(*sort)
    for order in orders:
        order.related = order.related_orders.all()
        if order.client.organization == '':
            order.client.organization_or_full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        else:
            order.client.organization_or_full_name = order.client.organization
        order.client.full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        prs = Order_Product.objects.filter(order_id=order.id, is_deleted=0)
        products_list = []
        for pr in prs:
            products_list.append(pr)
        order.products = products_list
        if order.bill_status == 0:
            order.bill_status = 'Выставлен'
        elif order.bill_status == 1:
            order.bill_status = 'Нужна доплата'
        elif order.bill_status == 2:
            order.bill_status = 'Оплачен'
        elif order.bill_status == 4:
            order.bill_status = 'Устно'
        elif order.bill_status == 5:
            order.bill_status = 'Подбор'
        else:
            order.bill_status = ''
        if order.bill != None:
            orders_count_str = str(order.bill)
            orders_count_str_reverse = orders_count_str[::-1]
            orders_count_str_right_format = ''
            for i in range(0, len(orders_count_str) / 3 + 1):
                j = 3
                if i != (len(orders_count_str) / 3):
                    orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[i * j] + \
                                                    orders_count_str_reverse[i * j + 1] + orders_count_str_reverse[
                                                        i * j + 2] + ' '
                else:
                    if (len(orders_count_str) % 3) == 2:
                        orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[
                            i * j] + \
                                                        orders_count_str_reverse[i * j + 1]
                    elif (len(orders_count_str) % 3) == 1:
                        orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[i * j]
            orders_count_str = orders_count_str_right_format[::-1]
            order.bill = orders_count_str
        order.files = []
        if Order_Files.objects.filter(order_id=order.id).all() is not None:
            for order_file in Order_Files.objects.filter(order_id=order.id).all():
                if order_file.file:
                    order_file.name = order_file.title
                    order_file.url = order_file.file.url
                    order.files.append(order_file)
    user_role = Roles.objects.get(id=request.user.id).role
    out.update({'user_role': user_role})
    out.update({'page_title': "Связанные заявки"})
    out.update({'related_claims': orders})
    out.update({'count': orders.count()})
    return render(request, 'order_claim/hidden/get_related_claims.html', out)


def get_client_claims(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    out.update({'sources': Sources.objects.filter(is_deleted=0)})
    out.update({'transport_campaigns': TransportCampaigns.objects.filter(is_deleted=0)})
    out.update({'roles': Roles.objects.filter(is_deleted=0).filter(Q(role=1) | Q(role=0)).all()})
    if 'page' in request.GET and 'length' in request.GET:
        page = int(request.GET['page'])
        length = int(request.GET['length'])
        start = (page - 1) * length
        out.update({'start': start})
    user_id = request.user.id
    out.update({'user_id': user_id})
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role == 2:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
    sort_key = request.GET.get('sort', DEFAULT_SORT_TYPE_FOR_CLAIM)
    sort = SORT_TYPE_FOR_CLAIM.get(sort_key, DEFAULT_SORT_TYPE_FOR_CLAIM)
    client = Clients.objects.get(id=request.GET['client-id'])
    claim = Orders.objects.get(id=request.GET['claim-id'])
    if client.organization == '':
        client.organization_or_full_name = client.last_name + ' ' + client.name + ' ' + client.patronymic
    else:
        client.organization_or_full_name = client.organization
    orders = Orders.objects.filter(is_deleted=0, is_claim=1, in_archive=0, client=client)
    if 'source' in request.GET:
        source = int(request.GET.get('source'))
        if source != -1:
            orders = orders.filter(source=source)
    if 'managers[]' in request.GET:
        managers = request.GET.getlist('managers[]')
        orders = orders.filter(role__in=managers)
        out.update({'managers': managers})
    try:
        orders = orders.order_by(sort)
    except TypeError:
        orders = orders.order_by(*sort)
    orders_list = [x for x in orders.all() if x not in claim.related_orders.all()]
    for order in orders_list:
        order.related = order.related_orders.all()
        if order.client.organization == '':
            order.client.organization_or_full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        else:
            order.client.organization_or_full_name = order.client.organization
        order.client.full_name = order.client.last_name + ' ' + order.client.name + ' ' + order.client.patronymic
        prs = Order_Product.objects.filter(order_id=order.id, is_deleted=0)
        products_list = []
        for pr in prs:
            products_list.append(pr)
        order.products = products_list
        if order.bill_status == 0:
            order.bill_status = 'Выставлен'
        elif order.bill_status == 1:
            order.bill_status = 'Нужна доплата'
        elif order.bill_status == 2:
            order.bill_status = 'Оплачен'
        elif order.bill_status == 4:
            order.bill_status = 'Устно'
        elif order.bill_status == 5:
            order.bill_status = 'Подбор'
        else:
            order.bill_status = ''
        if order.bill != None:
            orders_count_str = str(order.bill)
            orders_count_str_reverse = orders_count_str[::-1]
            orders_count_str_right_format = ''
            for i in range(0, len(orders_count_str) / 3 + 1):
                j = 3
                if i != (len(orders_count_str) / 3):
                    orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[i * j] + \
                                                    orders_count_str_reverse[i * j + 1] + orders_count_str_reverse[
                                                        i * j + 2] + ' '
                else:
                    if (len(orders_count_str) % 3) == 2:
                        orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[
                            i * j] + \
                                                        orders_count_str_reverse[i * j + 1]
                    elif (len(orders_count_str) % 3) == 1:
                        orders_count_str_right_format = orders_count_str_right_format + orders_count_str_reverse[i * j]
            orders_count_str = orders_count_str_right_format[::-1]
            order.bill = orders_count_str
        order.files = []
        if Order_Files.objects.filter(order_id=order.id).all() is not None:
            for order_file in Order_Files.objects.filter(order_id=order.id).all():
                if order_file.file:
                    order_file.name = order_file.title
                    order_file.url = order_file.file.url
                    order.files.append(order_file)
    user_role = Roles.objects.get(id=request.user.id).role
    out.update({'user_role': user_role})
    out.update({'page_title': "Заявки"})
    out.update({'claims': orders_list})
    out.update({'count': orders.count()})
    out.update({'organization_or_full_name': client.organization_or_full_name})
    return render(request, 'order_claim/hidden/get_client_claims.html', out)


def bind_claim(request):
    claim = Orders.objects.get(id=request.GET['id'])
    related_claim = Orders.objects.get(id=request.GET['related_with'])
    claim.related_orders.add(related_claim)
    claim.related_color = request.GET['color']
    related_claim.related_color = request.GET['color']
    related_claim.save(update_fields=["related_color"])
    claim.save(update_fields=["related_color"])
    return render(request, 'order_claim/hidden/get_client_claims.html')


def unbind_claim(request):
    claim = Orders.objects.get(id=request.GET['id'])
    related_claim = Orders.objects.get(id=request.GET['related_with'])
    claim.related_orders.remove(related_claim)
    return render(request, 'order_claim/hidden/get_client_claims.html')


def edit_template(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    if 'page' in request.GET and 'length' in request.GET:
        page = int(request.GET['page'])
        length = int(request.GET['length'])
        start = (page - 1) * length
        out.update({'start': start})
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role == 2 or user_role == 1:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
    out.update({'page_title': "Редактирование шаблона КП"})
    id = request.GET['id']
    template = KPTemplates.objects.filter(company__id=id).first()
    if request.method == 'POST':
        page = request.POST['page']
        kp_page = request.POST['kp-page']
        template_backup = KPTemplates.objects.create(html_text=page, html_text_for_kp=kp_page, number=id)
        if template:
            template.html_text = page
            template.html_text_for_kp = kp_page
        else:
            template = KPTemplates.objects.create(html_text=page, company=Companies.objects.get(pk=id),
                                                  html_text_for_kp=kp_page, number=1000)
        template.save()
    if template:
        out.update({'page': template.html_text})
        out.update({'kp_page': template.html_text_for_kp})
    return render(request, 'setting/edit_template.html', out)


def get_templates(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    if 'page' in request.GET and 'length' in request.GET:
        page = int(request.GET['page'])
        length = int(request.GET['length'])
        start = (page - 1) * length
        out.update({'start': start})
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role == 2 or user_role == 1:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
    out.update({'page_title': "Шаблоны КП"})
    companies = Companies.objects.filter(is_deleted=0)
    for c in companies:
        c.no_template = False
        c.number = u'Шаблон не создан'
        if KPTemplates.objects.filter(company=c).first():
            c.number = KPTemplates.objects.filter(company=c).first().number
            c.no_template = True
        c.full_name = c.last_name + ' ' + c.name + ' ' + c.patronymic
    out.update({'page_title': "Компании"})
    out.update({'companies': companies})
    out.update({'count': companies.count()})
    out.update({'form': KPTemplatesForm()})
    return render(request, 'setting/get_templates.html', out)


def edit_number_template(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role == 2 or user_role == 1:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
    get_params = '?'
    get_params += get_request_param_as_string(request)
    if request.method == 'POST':
        form = KPTemplatesForm(request.POST)
        if form.is_valid() and request.POST['number']:
            id = request.GET['id']
            kp_template = KPTemplates.objects.filter(company__id=id).first()
            number = request.POST['number']
            kp_template.number = number
            kp_template.save()
            return HttpResponseRedirect('/settings/template/' + get_params)
    return HttpResponseRedirect('/settings/template/' + get_params)


def edit_kp(request):
    if not request.user.is_active:
        return HttpResponseRedirect('/login/')
    out = {}
    if 'page' in request.GET and 'length' in request.GET:
        page = int(request.GET['page'])
        length = int(request.GET['length'])
        start = (page - 1) * length
        out.update({'start': start})
    user_role = Roles.objects.get(id=request.user.id).role
    if user_role == 2:
        return HttpResponseRedirect('/oops/')
    else:
        out.update({'user_role': user_role})
    out.update({'page_title': "Создание КП"})
    id = request.GET['id']
    claim = Orders.objects.get(pk=id)
    if request.method == 'GET':
        template = KPTemplates.objects.filter(company=claim.company).first()
        if not template:
            out.update({'page_title': "Для данной компании нет шаблона КП"})
            return render(request, 'edit_kp.html', out)
        temp_out = {}
        number = template.number
        template.number += 1
        template.save(update_fields=['number'])
        kp_date = date.today().strftime('%d.%m.%Y')
        products = Order_Product.objects.filter(order_id=claim.id, is_deleted=0).all()
        count = 1
        for product in products:
            product.number = count
            count += 1
            product.total_price = product.count_of_products * product.price
            product.title = product.product.title
        if claim.client.organization == '':
            organization_or_full_name = claim.client.last_name + ' ' + claim.client.name + ' ' + claim.client.patronymic
        else:
            organization_or_full_name = claim.client.organization
        # TODO
        added_table = ''
        temp_out.update({'number': number})
        account_number = 'КП' + str(number)
        claim.account_number = account_number
        claim.save(update_fields=["account_number"])
        temp_out.update({'date': kp_date})
        temp_out.update({
            'organization_name': u'<input type="text" class="organization_name" name="organization_name" value="{}">'.format(
                organization_or_full_name)})
        temp_out.update({'added_table': added_table})
        temp_out.update({'products': products})
        temp_out.update({'in_total': claim.bill})
        temp_out.update({'manager': claim.role})
        html = Template(template.html_text_for_kp.encode('utf-8')).render(Context(temp_out))
        out.update({'page': html})
        return render(request, 'edit_kp.html', out)
    else:
        temp_out = {}
        organization_name = request.POST.get('organization_name', '')
        accompanying_text = request.POST.get('accompanying_text', '')
        table__total_kp = request.POST.get('table__total_kp', '')
        table__total = request.POST.get('table__total', '')
        added_table = ''
        added_table_kp = ''
        if 'added_table' in request.POST:
            added_table__total_kp = request.POST.get('added_table__total_kp', '')
            added_table__total = request.POST.get('added_table__total', '')
            added_table = request.POST.get('added_table', '')
            added_table_kp = request.POST.get('added_table', '')
            added_table_out = {}
            added_table_kp_out = {}
            added_table_out.update({'added_table__total': added_table__total})
            added_table_kp_out.update({'added_table__total': added_table__total_kp})
            added_table = Template(added_table).render(Context(added_table_out))
            added_table_kp = Template(added_table_kp).render(Context(added_table_kp_out))
        page = request.POST.get('page', '')
        temp_out.update({'accompanying_text': accompanying_text})
        temp_out.update({'added_table': added_table_kp})
        temp_out.update({'table__total': table__total_kp})
        temp_out.update({'organization_name': organization_name})
        html = Template(page).render(Context(temp_out))
        page_out = {}
        page_out.update({'accompanying_text': accompanying_text})
        page_out.update({'added_table': added_table})
        page_out.update({'table__total': table__total})
        page_out.update({
            'organization_name': u'<input type="text" class="organization_name" name="organization_name" value="{}">'.format(
                organization_name)})
        page_html = Template(page).render(Context(page_out))
        out.update({'page': page_html})
        symbols = (u"абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ ",
                   u"abvgdeejzijklmnoprstufhzcss_y_euaABVGDEEJZIJKLMNOPRSTUFHZCSS_Y_EUA_")
        tr = dict( [ (ord(a), ord(b)) for (a, b) in zip(*symbols) ] )
        filename = str(claim.account_number.translate(tr)) + '_' + str(claim.company.title.translate(tr)) + '_' + str(hash(datetime.now()))
        out_filename_pdf = MEDIA_ROOT + '/' + str('uploads/') + filename + '.pdf'
        out_filename_html = MEDIA_ROOT + '/' + str('uploads/') + filename + '.html'
        form_file = open(out_filename_html, 'wb')
        form_file.write(html.encode('utf-8'))
        form_file.close()
        out.update({'filename': filename})

        options = {
            'page-size': 'Letter',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
        }
        pdfkit.from_file(out_filename_html,
                         out_filename_pdf,
                         options=options)
        return render(request, 'edit_kp.html', out)
